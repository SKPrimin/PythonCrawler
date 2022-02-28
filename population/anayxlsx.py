import csv
import re

from openpyxl import load_workbook


def getSheet(sheetName,header,before,after):
    '''表格转换函数,四参数：sheet名称,首行名称，前索引，后索引'''
    try:
        # 获取指定的表单
        ws = wb[sheetName]
        # 正则表达式匹配出表格名
        filename = re.sub(r'\d-\d+','',sheetName)
        # 创建对应的.csv文件，从3或4开始切片，追加模式，utf-8编码,新建""一行
        with open(f'{filename}.csv', mode="a", encoding='utf-8', newline="") as f:
            # 创建filepencil，用来在问价上写入数据
            fp = csv.writer(f)
            # 写入表头
            fp.writerow(header)
            for index in range(before,after+1):
                rowitems = []
                for index,item in enumerate(ws[index]):
                    rowitem = item.value if index != 0 else item.value.replace(' ', '')
                    rowitems.append(rowitem)
                fp.writerow(rowitems)
    except Exception as e:
        print(e)


if __name__ == '__main__':
    # 加载表格文件
    wb = load_workbook('ahpopulation.xlsx')
    # # 主要年份人口指标
    # header31 = ['年份', '户籍人口总数（万人）', '城镇人口比重%', '常住人口总数（万人）', '城镇人口比重%',
    #  '出生率‰', '死亡率‰', '自然增长率‰', '流向省外半年以上的人数（万人）']
    # getSheet(sheetName='3―1主要年份人口指标', header=header31, before=5, after=15)

    # 主要年份人口系数
    header32 = ['年份', '少年儿童系数', '老年系数', '老少比', '少年儿童抚养系数',
                '老年抚养系数', '总抚养系数', '年龄中位数（岁）']
    getSheet(sheetName='3―2主要年份人口系数', header=header32, before=3, after=13)

    # # 各市主要年份人口城镇化率
    # header32 = ['地区', '2010', '2015', '2017', '2018', '2019']
    # getSheet(sheetName='3―4各市主要年份人口城镇化率', header=header32, before=1, after=17)

    # # 3―5各市常住人口出生率、死亡率（2019年）
    # header35 = ['地区', '出生率（‰）', '死亡率（‰）', '自然增长率（‰）']
    # getSheet(sheetName='3―5各市常住人口出生率、死亡率（2019年）', header=header35, before=1, after=17)

    # # 3―7各市户数、人口数和性别比（2019年）
    # header37 = ['地区（万户）', '户数（万户）', '人口数', '男', '性别比（女=100）']
    # getSheet(sheetName='3―7各市户数、人口数和性别比（2019年）', header=header37, before=1, after=17)

    # # 3―10按年龄和性别分人口数（2019年）
    # header310 = ['年龄', '人口总数（人）', '男性总数', '女性总数', '占总人口比重（%）', '男性比重', '女性比重', '性别比（女=100）']
    # getSheet(sheetName='3―10按年龄和性别分人口数（2019年）', header=header310, before=1, after=15)

    # # 3―12各市按家庭户规模分的户数构成（2019年）
    # header312 = ['地区', '家庭户规模（人/户）', '一人户', '二人户', '三人户', '四人户', '五人户', '六人及六人以上户（人/户）']
    # getSheet(sheetName='3―12各市按家庭户规模分的户数构成（2019年）', header=header312, before=2, after=18)

    # # 3―21各市流向省外半年以上的流动人口构成（2019年）
    # header321 = ['地区', '合计', '江苏', '浙江', '上海', '广东', '北京', '福建', '山东', '天津', '河南', '河北', '新疆', '辽宁', '湖北', '陕西', '流向其他省市']
    # getSheet(sheetName='3―21各市流向省外半年以上的流动人口构成（2019年）', header=header321, before=2, after=18)

    # # 3―25历 年 全 省 总 人 口、总 户 数
    # header325 = ['年份','总户数','合计总人口','男性人口','女性人口','性别比(女=100)','城镇人口','乡村人口']
    # getSheet(sheetName='3―25历 年 全 省 总 人 口、总 户 数', header=header325, before=2, after=17)

    # # 3―26各市、县、区户数、人口数（2019年）
    # header326 = ['地区', '总户数（万人）', '户籍人口（万人）', '男户籍人口', '女户籍人口', '性别比（女=100）', '城镇人口', '常住人口']
    # getSheet(sheetName='3―26各市、县、区户数、人口数（2019年）', header=header326, before=1, after=137)
